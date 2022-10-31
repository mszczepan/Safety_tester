import openpyxl as op
from openpyxl.utils import get_column_letter


# Loading and processing of the safety matrix


class ExcelService:

    # Initialization of the safety matrix coordinates

    def __init__(self, row_start, row_end, column_start, column_end):
        self.sh = None
        self.ws = None
        self.wb = None
        self.row_start = row_start
        self.row_end = row_end
        self.column_start = column_start
        self.column_end = column_end
        self.matrix_conf = {}

    def load_wb(self, path, sheet_name):
        self.wb = op.load_workbook(path)
        self.sh = self.wb[sheet_name]

    def read_matrix(self):

        # building of matrix structure

        for row in range(self.row_start, self.row_end):
            self.matrix_conf[self.sh[f'{get_column_letter(self.column_start)}{row + 1}'].value] = []

            for column in range(self.column_start + 1, self.column_end + 1):
                self.matrix_conf[self.sh[f'{get_column_letter(self.column_start)}{row + 1}'].value].append(
                    {self.sh[f'{get_column_letter(column)}{self.row_start}'].value: None})

        # matrix filling

        row_number = self.row_start + 1

        for key, value in self.matrix_conf.items():
            count = 0
            column_number = self.column_start + 1
            for item in value:
                if self.sh[f'{get_column_letter(column_number)}{row_number}'].value is None:
                    self.matrix_conf[key][count][list(item.keys())[0]] = 0
                elif self.sh[f'{get_column_letter(column_number)}{row_number}'].value == 'x':
                    self.matrix_conf[key][count][list(item.keys())[0]] = 1
                column_number = column_number + 1
                count = count + 1
            row_number = row_number + 1

            return self.matrix_conf

